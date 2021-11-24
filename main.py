from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update, ReplyKeyboardMarkup
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, CallbackQueryHandler, CallbackContext
from boto.s3.connection import S3Connection
import telegram
import random
import string
import requests
import logging
import json
import os
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from random import randint, randrange

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

TOKEN = os.getenv("TOKEN")
# guzo bus main bot
# 565110335
ADMIN = [344049097, 2052373171, 565110335, 1799128648]


# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

logger = logging.getLogger(__name__)


class GuzoBusBot:
    def __init__(self, TOKEN):
        self.TOKEN = TOKEN
        self.updater = Updater(TOKEN, use_context=True)
        self.dispatcher = self.updater.dispatcher
        self.bot = telegram.Bot(self.TOKEN)

        self.messages = dict()
        self.english_messages = self.load_data()
        self.amharic_messages = self.load_data(False)
        self.user_inputs = dict()
        self.user_status = dict()
        self.user_section = dict()

    def load_data(self, is_en=True, switch=False, chat_id=0):
        if switch:
            is_en = not is_en
        if is_en:
            lang = "en"
        else:
            lang = "am"
        with open(f"{CURRENT_DIR}/data/langs/{lang}.json", "rb") as f:
            data = json.load(f)
            f.close()
        if not switch:
            return data
        else:
            self.messages[chat_id] = lang

    def handle_image(self, update, ctx):
        chat_id = update.effective_chat.id

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages

        ckeyboard = [
            [InlineKeyboardButton(messages["back_to_main"],
                                  callback_data="back_main")]
        ]
        creply_markup = InlineKeyboardMarkup(ckeyboard)

        if self.user_status[chat_id] == "upload_proof_image" and len(self.user_inputs[chat_id]) > 1:
            try:
                # validates whether the pnr is valid or not
                valid = False
                print(update.message.photo[-1])
                file = ctx.bot.getFile(update.message.photo[-1].file_id)

                # file.download("image.jpg")
                valid = True
                # if the message is valid then the bot will ask for a   message
                if valid:
                    reply_markup = ReplyKeyboardMarkup(
                        [[messages["cancel"], messages["submit"]]], resize_keyboard=True, one_time_keyboard=True)
                    ctx.bot.send_message(chat_id, messages["confirm"])
                    ctx.bot.send_message(
                        chat_id, messages["phone"]+": "+self.user_inputs[chat_id][1]+"\n" + "PNR: "+self.user_inputs[chat_id][0], reply_markup=reply_markup)
                    self.user_inputs[chat_id].append(file)
                    self.user_status[chat_id] = "booking_confirmation"

                # if the message is invalid then the bot will redirect  user back to passenger menu
                else:
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])

            except Exception as e:
                print(e)
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])
        elif self.user_status[chat_id] == "upload_weekly_file" and len(self.user_inputs[chat_id]) > 4:
            ctx.bot.send_message(
                chat_id, messages["send_file"], reply_markup=creply_markup)

    def handle_document(self, update, ctx):
        chat_id = update.effective_chat.id

        try:
            if(chat_id in ADMIN and update.message.document.file_name == "seats.xlsx"):
                ctx.bot.getFile(update.message.document).download(
                    "data/seats.xlsx")
                ctx.bot.send_message(chat_id, "Seats Successfully Updated")
            elif(chat_id in ADMIN and update.message.document.file_name == "operators.xlsx"):
                ctx.bot.getFile(update.message.document).download(
                    "data/operators.xlsx")
                ctx.bot.send_message(
                    chat_id, "Operators Successfully Updated")
            elif(chat_id in ADMIN and update.message.document.file_name == "buses.txt"):
                ctx.bot.getFile(update.message.document).download(
                    "data/buses.txt")
                ctx.bot.send_message(chat_id, "Buses Successfully Updated")
            elif(chat_id in ADMIN and update.message.document.file_name == "bookings.xlsx"):
                ctx.bot.getFile(update.message.document).download(
                    "data/bookings.xlsx")
                ctx.bot.send_message(chat_id, "Bookings Successfully Updated")
            elif(chat_id in ADMIN and update.message.document.file_name == "locations.txt"):
                ctx.bot.getFile(update.message.document).download(
                    "data/locations.txt")
                ctx.bot.send_message(
                    chat_id, "Locations Successfully Updated")
            elif(chat_id in ADMIN and update.message.document.file_name == "en.json"):
                ctx.bot.getFile(update.message.document).download(
                    "data/langs/en.json")
                ctx.bot.send_message(
                    chat_id, "English Successfully Updated")
                self.english_messages = self.load_data()
                self.amharic_messages = self.load_data(False)
            elif(chat_id in ADMIN and update.message.document.file_name == "am.json"):
                ctx.bot.getFile(update.message.document).download(
                    "data/langs/am.json")
                ctx.bot.send_message(
                    chat_id, "Amharic Successfully Updated")
                self.english_messages = self.load_data()
                self.amharic_messages = self.load_data(False)

            if self.messages[chat_id] == "en":
                messages = self.english_messages
            else:
                messages = self.amharic_messages

            ckeyboard = [
                [InlineKeyboardButton(messages["back_to_main"],
                                      callback_data="back_main")]]
            creply_markup = InlineKeyboardMarkup(ckeyboard)

            if self.user_status[chat_id] == "upload_weekly_file" and len(self.user_inputs[chat_id]) > 4:
                try:
                    # validates whether the pnr is valid or not
                    valid = False
                    file = ctx.bot.getFile(update.message.document)
                    valid = True
                    # if the message is valid then the bot will ask for a   message
                    if valid:
                        reply_markup = ReplyKeyboardMarkup(
                            [[messages["cancel"], messages["submit"]]], resize_keyboard=True, one_time_keyboard=True)
                        ctx.bot.send_message(chat_id, self.user_inputs[chat_id][0]
                                             + ", " +
                                             messages["file_confirmation"],
                                             reply_markup=reply_markup)
                        self.user_inputs[chat_id].append(file)
                        self.user_status[chat_id] = "weekly_trip_confirmation"

                    # if the message is invalid then the bot will redirect  user back to  passenger menu
                    else:
                        ctx.bot.send_message(chat_id, messages["wrong_value"])
                        ctx.bot.send_message(chat_id, messages["try_again"])

                except Exception as e:
                    print(e)
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
            elif self.user_status[chat_id] == "upload_proof_image" and len(self.user_inputs[chat_id]) > 1:
                ctx.bot.send_message(
                    chat_id, messages["send_image"], reply_markup=creply_markup)
        except Exception as e:
            print(e)

    def handle_text(self, update, ctx):
        chat_id = update.effective_chat.id
        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages
        if update.message.text == "English 🇬🇧":
            self.load_data(False, True, chat_id)
            self.main_menu(update, ctx)
        elif update.message.text == "አማርኛ 🇪🇹":
            self.load_data(True, True, chat_id)
            self.main_menu(update, ctx)
        elif chat_id in ADMIN and update.message.text.lower() == "admin":
            self.admin_menu(update, ctx)
        elif chat_id in ADMIN and update.message.text.lower() == "add pnr":
            self.add_pnr(update, ctx)
        elif chat_id in ADMIN and update.message.text.lower() == "seats":
            files = {"document": open(f"data/seats.xlsx", 'rb')}
            resp = requests.post("https://api.telegram.org/bot" +
                                 TOKEN+"/sendDocument?chat_id="+str(chat_id), files=files)
            print(resp.status_code)
        elif chat_id in ADMIN and update.message.text.lower() == "bookings":
            files = {"document": open(f"data/bookings.xlsx", 'rb')}
            resp = requests.post("https://api.telegram.org/bot" +
                                 TOKEN+"/sendDocument?chat_id="+str(chat_id), files=files)
            print(resp.status_code)
        elif chat_id in ADMIN and update.message.text.lower() == "locations":
            files = {"document": open(f"data/locations.txt", 'rb')}
            resp = requests.post("https://api.telegram.org/bot" +
                                 TOKEN+"/sendDocument?chat_id="+str(chat_id), files=files)
            print(resp.status_code)
        elif chat_id in ADMIN and update.message.text.lower() == "operators":
            files = {"document": open(f"data/operators.xlsx", 'rb')}
            resp = requests.post("https://api.telegram.org/bot" +
                                 TOKEN+"/sendDocument?chat_id="+str(chat_id), files=files)
            print(resp.status_code)
        elif chat_id in ADMIN and update.message.text.lower() == "operators":
            files = {"document": open(f"data/operators.xlsx", 'rb')}
            resp = requests.post("https://api.telegram.org/bot" +
                                 TOKEN+"/sendDocument?chat_id="+str(chat_id), files=files)
            print(resp.status_code)
        elif chat_id in ADMIN and update.message.text.lower() == "english":
            files = {"document": open(f"data/langs/en.json", 'rb')}
            resp = requests.post("https://api.telegram.org/bot" +
                                 TOKEN+"/sendDocument?chat_id="+str(chat_id), files=files)
            print(resp.status_code)
        elif chat_id in ADMIN and update.message.text.lower() == "amharic":
            files = {"document": open(f"data/langs/am.json", 'rb')}
            resp = requests.post("https://api.telegram.org/bot" +
                                 TOKEN+"/sendDocument?chat_id="+str(chat_id), files=files)
            print(resp.status_code)
        elif update.message.text == messages["back"]:
            self.main_menu(update, ctx)
        elif update.message.text == messages["referal_link"]:
            self.refer(update, ctx)
        elif update.message.text == messages["about_us"]:
            self.about(update, ctx)
        elif update.message.text == messages["rules"]:
            self.rules(update, ctx)
        elif update.message.text == messages["contact_us"]:
            self.contactus(update, ctx)
        elif update.message.text == messages["language_switcher"]:
            self.start(update, ctx)
        elif update.message.text == messages["quest_operator"]:
            self.operator_menu(update, ctx)
        elif update.message.text == messages["quest_passenger"]:
            self.passenger_menu(update, ctx)
        elif update.message.text == messages["book_bus"]:
            self.book_bus(update, ctx)
        elif update.message.text == messages["proof"]:
            self.upload_proof(update, ctx)
        elif update.message.text == messages["sell_seats"]:
            self.sell_seats(update, ctx)
        elif update.message.text == messages["upload_list"]:
            self.upload_weekly_trip(update, ctx)
        elif update.message.text == messages["verify_passenger"]:
            self.verify_passenger(update, ctx)
        elif chat_id in self.user_section:
            if self.user_section[chat_id] == "book_bus":
                self.book_bus_process(update, ctx, chat_id, messages)
            elif self.user_section[chat_id] == "payment_proof":
                self.upload_proof_process(update, ctx, chat_id, messages)
            elif self.user_section[chat_id] == "sell_seats":
                self.sell_seats_process(update, ctx, chat_id, messages)
            elif self.user_section[chat_id] == "upload_trip":
                self.upload_trip_process(update, ctx, chat_id, messages)
            elif self.user_section[chat_id] == "verify_passenger":
                self.verify_passenger_process(update, ctx, chat_id, messages)
            elif self.user_section[chat_id] == "add_pnr":
                self.add_pnr_process(update, ctx, chat_id, messages)

    # handles the process of proving the payment

    def upload_proof_process(self, update, ctx, chat_id, messages):
        ckeyboard = [
            [InlineKeyboardButton(messages["back_to_main"],
                                  callback_data="back_main")]
        ]
        creply_markup = InlineKeyboardMarkup(ckeyboard)
        # checks if user is in select pnr step
        if self.user_status[chat_id] == "enter_pnr":
            # validates whether the pnr is valid or not
            valid = False
            if(len(update.message.text) == 8):
                valid = True
            # if the pnr is valid then the bot will ask for a message
            if valid:
                ctx.bot.send_message(
                    chat_id, messages["enter_phone"], reply_markup=creply_markup)
                self.user_status[chat_id] = "enter_phone"
                self.user_inputs[chat_id] = [update.message.text]
            # if the pnr is invalid then the bot will redirect user back to passenger menu
            else:
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])
                self.user_status.pop(chat_id, None)
                self.user_section.pop(chat_id, None)
                self.user_inputs.pop(chat_id, None)
                self.passenger_menu(update, ctx)

        # checks if user is in enter name step
        elif self.user_status[chat_id] == "enter_phone" and len(self.user_inputs[chat_id]) > 0:
            try:
                valid = False
                if(len(update.message.text) == 10):
                    if(update.message.text[0:1] == "0" and update.message.text.isdigit()):
                        valid = True
                if valid:
                    if self.check_pnr(update.message.text, self.user_inputs[chat_id][0]) == True:
                        ctx.bot.send_message(
                            chat_id, messages["upload_proof_image"],  reply_markup=creply_markup)
                        self.user_status[chat_id] = "upload_proof_image"
                        self.user_inputs[chat_id].append(update.message.text)
                    else:
                        ctx.bot.send_message(
                            chat_id, messages["wrong_pnr_credentials"])
                        ctx.bot.send_message(chat_id, messages["try_again"])
                        self.user_status.pop(chat_id, None)
                        self.user_section.pop(chat_id, None)
                        self.user_inputs.pop(chat_id, None)
                        self.passenger_menu(update, ctx)
                else:
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
            except Exception as e:
                print(e)
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])

        # checks if user is in select message step
        elif self.user_status[chat_id] == "booking_confirmation" and len(self.user_inputs[chat_id]) > 1:
            try:
                # validates whether the confirmation is submitted or not
                if(update.message.text == messages["submit"]):
                    # if the user submits the confirmation then the system will send the data to the admin and delete the picture
                    filename = self.user_inputs[chat_id][2].file_path.split(
                        "/")[-1]
                    self.user_inputs[chat_id][2].download("data/"+filename)
                    print(self.user_inputs[chat_id])
                    dphone = self.user_inputs[chat_id][1]
                    dpnr = self.user_inputs[chat_id][0]
                    ctx.bot.send_message(
                        chat_id, messages["payment_proof_submitted"], reply_markup=creply_markup)
                    for adm in ADMIN:
                        files = {"photo": open(f"data/{filename}", 'rb')}
                        ctx.bot.send_message(adm, messages["new_payment_proof"] +
                                             #   +"\n@"+update.message.from_user["username"]
                                             "\n"+messages["phone"] + \
                                             ": 0"+str(int(dphone))
                                             + "\nPNR: "+dpnr+"\n")
                        resp = requests.post(
                            "https://api.telegram.org/bot"+TOKEN+"/sendPhoto?chat_id="+str(adm), files=files)
                        print(resp)
                    self.user_status.pop(chat_id, None)
                    self.user_section.pop(chat_id, None)
                    self.user_inputs.pop(chat_id, None)

                    os.remove(f"{CURRENT_DIR}/data/{filename}")

                else:
                    self.user_status.pop(chat_id, None)
                    self.user_section.pop(chat_id, None)
                    self.user_inputs.pop(chat_id, None)
                    self.passenger_menu(update, ctx)
                valid = False
            except Exception as e:
                print("exception")
                if str(e) == "Chat not found":
                    print("Chat not fouond")
                elif str(e) == "Forbidden: bot was blocked by the user":
                    print("Bot Blocked by user")
                else:
                    print(e)
                    # if the date has an error then the bot will redirect user back to passenger menu
                    #ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
        elif self.user_status[chat_id] == "upload_proof_image":
            # if the user sends message instead of image, the system will ask again
            ctx.bot.send_message(
                chat_id, messages["send_image"], reply_markup=creply_markup)

    # handles the process of booking a bus

    def book_bus_process(self, update, ctx, chat_id, messages):
        ckeyboard = [
            [InlineKeyboardButton(messages["back_to_main"],
                                  callback_data="back_main")]
        ]
        creply_markup = InlineKeyboardMarkup(ckeyboard)
        # checks if user is in select source step
        if self.user_status[chat_id] == "select_source":
            locations = []
            with open(f"{CURRENT_DIR}/data/locations.txt", "rb") as f:
                for line in f:
                    locations.append([line.decode().rstrip()])
                locations.append([messages["back"]])
                f.close()

                # validates whether the source is valid or not
                valid = False
                for line in open(f"{CURRENT_DIR}/data/locations.txt", "rb"):
                    if line.decode().rstrip() == update.message.text:
                        self.user_inputs[chat_id] = [update.message.text]
                        valid = True
                        break

                # if the source is valid then the bot will ask for destination
                if valid:
                    ctx.bot.send_message(chat_id, messages["select_destination"],
                                         reply_markup=ReplyKeyboardMarkup(locations, one_time_keyboard=True))

                    self.user_status[chat_id] = "select_destination"

                # if the source is invalid then the bot will redirect user back to passenger menu
                else:
                    ctx.bot.send_message(chat_id, messages["wrong_location"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
                    self.user_status.pop(chat_id, None)
                    self.user_section.pop(chat_id, None)
                    self.user_inputs.pop(chat_id, None)
                    self.passenger_menu(update, ctx)

        # checks if user is in select destination step
        elif self.user_status[chat_id] == "select_destination" and len(self.user_inputs[chat_id]) > 0:
            # validates whether the destination is valid or not
            valid = False
            for line in open(f"{CURRENT_DIR}/data/locations.txt", "rb"):
                if line.decode().rstrip() == update.message.text:
                    self.user_inputs[chat_id].append(update.message.text)
                    valid = True
                    break

            # if the destination is valid then the bot will ask for trip date
            if valid:
                ctx.bot.send_message(
                    chat_id, messages["select_trip_date"], reply_markup=creply_markup)
                self.user_status[chat_id] = "select_trip_date"

            # if the destination is invalid then the bot will redirect user back to passenger menu
            else:
                ctx.bot.send_message(chat_id, messages["wrong_location"])
                ctx.bot.send_message(chat_id, messages["try_again"])
                self.user_status.pop(chat_id, None)
                self.user_section.pop(chat_id, None)
                self.user_inputs.pop(chat_id, None)
                self.passenger_menu(update, ctx)

        # checks if user is in select trip date step
        elif self.user_status[chat_id] == "select_trip_date" and len(self.user_inputs[chat_id]) > 1:
            try:
                # validates whether the date is valid or not
                if self.date_validator(update.message.text):
                    # if the date is valid then the bot will ask for passenger number
                    ctx.bot.send_message(
                        chat_id, messages["enter_number_of_passengers"], reply_markup=creply_markup)
                    self.user_status[chat_id] = "enter_number_of_passengers"
                    self.user_inputs[chat_id].append(update.message.text)

                # if the date is invalid then the bot will redirect user back to passenger menu
                else:
                    ctx.bot.send_message(chat_id, messages["wrong_date"])
                    ctx.bot.send_message(chat_id, messages["try_again"])

            except Exception as e:
                print(e)
                # if the date has an error then the bot will redirect user back to passenger menu
                ctx.bot.send_message(chat_id, messages["wrong_date"])
                ctx.bot.send_message(chat_id, messages["try_again"])

        # checks if user is in enter number of passengers step
        elif self.user_status[chat_id] == "enter_number_of_passengers" and len(self.user_inputs[chat_id]) > 2:
            try:
                # validates whether the number is valid or not
                valid = False
                if(update.message.text.isdigit()):
                    valid = True

                   # if the date is valid then the bot will check the excel file
                if valid:
                    self.user_inputs[chat_id].append(update.message.text)
                    self.user_status[chat_id] = "confirmation"
                    nbuses = self.seat_match_maker(
                        self.user_inputs[chat_id][2], self.user_inputs[chat_id][0], self.user_inputs[chat_id][1], self.user_inputs[chat_id][3])
                    keyboard = []
                    for avbus in nbuses:
                        keyboard.append([avbus])
                    keyboard.append([messages["back"]])
                    if(len(nbuses) > 0):
                        ctx.bot.send_message(chat_id, messages["select_bus"],
                                             reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))

                        self.user_status[chat_id] = "select_bus"
                    else:
                        keyboard = [[InlineKeyboardButton(
                            messages["back_to_main"], callback_data="back_main")]]
                        reply_markup = InlineKeyboardMarkup(keyboard)
                        ctx.bot.send_message(chat_id, messages["for"]+" "+messages["source"]+" : "+self.user_inputs[chat_id][0] +
                                             " - "+messages["destination"]+" : "+self.user_inputs[chat_id][1]+" "+messages["no_buses_found"], reply_markup=reply_markup)
                        self.user_status.pop(chat_id, None)
                        self.user_section.pop(chat_id, None)
                        self.user_inputs.pop(chat_id, None)
                    # reply_markup = ReplyKeyboardMarkup([[messages["cancel"], messages["submit"]]],resize_keyboard=True)
                    # ctx.bot.send_message(chat_id, "Confrimation", reply_markup = reply_markup)
                # if the number is invalid then the bot will redirect user back to passenger menu
                else:
                    ctx.bot.send_message(chat_id, messages["wrong_number"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
                    self.user_status.pop(chat_id, None)
                    self.user_section.pop(chat_id, None)
                    self.user_inputs.pop(chat_id, None)
                    self.passenger_menu(update, ctx)
            except Exception as e:
                print(e)
                # if the number has an error then the bot will redirect user back to passenger menu
                ctx.bot.send_message(chat_id, messages["wrong_number"])
                ctx.bot.send_message(chat_id, messages["try_again"])
                self.user_status.pop(chat_id, None)
                self.user_section.pop(chat_id, None)
                self.user_inputs.pop(chat_id, None)
                self.passenger_menu(update, ctx)

        # checks if user is in select bus step
        elif self.user_status[chat_id] == "select_bus" and len(self.user_inputs[chat_id]) > 3:
            ctx.bot.send_message(
                chat_id, messages["enter_name"], reply_markup=creply_markup)
            self.user_status[chat_id] = "enter_name"
            self.user_inputs[chat_id].append(update.message.text)

        # checks if user is in enter name step
        elif self.user_status[chat_id] == "enter_name" and len(self.user_inputs[chat_id]) > 4:
            ctx.bot.send_message(
                chat_id, messages["enter_phone"], reply_markup=creply_markup)
            self.user_status[chat_id] = "enter_phone"
            self.user_inputs[chat_id].append(update.message.text)

        # checks if user is in enter phone step
        elif self.user_status[chat_id] == "enter_phone" and len(self.user_inputs[chat_id]) > 5:
            try:
                valid = False
                if(len(update.message.text) == 10):
                    if(update.message.text[0:1] == "0" and update.message.text.isdigit()):
                        valid = True
                # if the pnr is valid then the bot will ask for a code
                if valid:
                    self.user_inputs[chat_id].append(update.message.text)
                    busprice = self.seat_reserver(self.user_inputs[chat_id][2], self.user_inputs[chat_id][0],
                                                  self.user_inputs[chat_id][1], self.user_inputs[chat_id][3], self.user_inputs[chat_id][4])
                    randid = randint(10000000, 99999999)
                    if busprice > 0:
                        # bus_book_follow_up
                        self.add_bookings_to_excel([self.user_inputs[chat_id][2], self.user_inputs[chat_id][5],
                                                    self.user_inputs[chat_id][6],
                                                    self.user_inputs[chat_id][0],
                                                    self.user_inputs[chat_id][1],
                                                    self.user_inputs[chat_id][2],
                                                    self.user_inputs[chat_id][3],
                                                    self.user_inputs[chat_id][4],
                                                    busprice,
                                                    int(int(
                                                        self.user_inputs[chat_id][3]) * int(busprice)),
                                                    randid
                                                    ])
                        keyboard = [
                            [InlineKeyboardButton(
                                messages["back_to_main"], callback_data="back_main")]
                        ]
                        reply_markup = InlineKeyboardMarkup(keyboard)
                        dname = self.user_inputs[chat_id][5]
                        dphone = self.user_inputs[chat_id][6]
                        ddate = self.user_inputs[chat_id][2]
                        dsource = self.user_inputs[chat_id][0]
                        ddestination = self.user_inputs[chat_id][1]
                        dseats = self.user_inputs[chat_id][3]
                        dbus = self.user_inputs[chat_id][4]
                        dprice = busprice
                        str(int(
                            int(self.user_inputs[chat_id][3]) * int(busprice)))
                        ctx.bot.send_message(chat_id, messages["bus_book_follow_front"]+"\n" +
                                             messages["date"]+" : "+self.user_inputs[chat_id][2] +
                                             "\n"+messages["source"]+" : "+self.user_inputs[chat_id][0] +
                                             "\n"+messages["destination"]+" : "+self.user_inputs[chat_id][1] +
                                             "\n"+messages["seats"]+" : "+self.user_inputs[chat_id][3] +
                                             "\n"+messages["bus"]+" : "+self.user_inputs[chat_id][4] +
                                             "\n\n"+messages["bus_book_follow_up"] % str(int(int(self.user_inputs[chat_id][3]) * int(busprice))), reply_markup=reply_markup)

                        for adm in ADMIN:
                            ctx.bot.send_message(adm, messages["new_bus_book_checking"]
                                                 + "\n"+"ID : "+str(randid)
                                                 + "\n" +
                                                 messages["name"]+" : "+dname
                                                 + "\n" +
                                                 messages["phone"]+" : "+dphone
                                                 + "\n" +
                                                 messages["source"] +
                                                 " : "+dsource
                                                 + "\n"+messages["destination"] +
                                                 " : " + ddestination
                                                 + "\n" +
                                                 messages["date"] +
                                                 " : " + ddate
                                                 + "\n" +
                                                 messages["seats"] +
                                                 " : " + dseats
                                                 + "\n" +
                                                 messages["bus"]+" : " + dbus
                                                 + "\n"+messages["price"] +
                                                 " : " + str(int(dprice))
                                                 + "\n"+messages["total_price"]+" : " +
                                                 str(int(int(dseats) * int(dprice)))
                                                 )
                            files = {"document": open(
                                f"data/seats.xlsx", 'rb')}
                            resp = requests.post("https://api.telegram.org/bot" +
                                                 TOKEN+"/sendDocument?chat_id="+str(adm), files=files)
                            print(str(adm))
                            print("seats")
                            print(resp)
                            files = {"document": open(
                                f"data/bookings.xlsx", 'rb')}
                            resp = requests.post("https://api.telegram.org/bot" +
                                                 TOKEN+"/sendDocument?chat_id="+str(adm), files=files)
                            print(str(adm))
                            print("bookings")
                            print(resp)

                            # "\n@"+update.message.from_user["username"]+"\n"+messages["source"]+" : "+self.user_inputs[chat_id][0]+"\n"+messages["destination"]+" : " +self.user_inputs[chat_id][1]+"\n"+messages["date"]+" : "+self.user_inputs[chat_id][2]+"\n"+messages["seats"]+" : "+self.user_inputs[chat_id][3]+"\n"+messages["bus"]+" : "+self.user_inputs[chat_id][4])
                            #   +"\n"+messages["price"]+" : "+str(int(busprice))+"\n"+messages["total_price"]+" : "+str(int(int(self.user_inputs[chat_id][3]) *int(busprice)))

                    else:
                        print("1st else")
                        ctx.bot.send_message(
                            chat_id, messages["booking_error"])
                        self.user_status.pop(chat_id, None)
                        self.user_section.pop(chat_id, None)
                        self.user_inputs.pop(chat_id, None)
                        self.passenger_menu(update, ctx)

                # if the phone is invalid then the bot will ask again
                else:
                    print("2nd else")
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])

            except Exception as e:
                print("exception")
                if str(e) == "Chat not found":
                    print("Chat not fouond")
                elif str(e) == "Forbidden: bot was blocked by the user":
                    print("Bot Blocked by user")
                else:
                    print(e)
                    # if the date has an error then the bot will redirect user back to passenger menu
                    #ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])

    # handles the process of selling seats

    def sell_seats_process(self, update, ctx, chat_id, messages):
        ckeyboard = [
            [InlineKeyboardButton(messages["back_to_main"],
                                  callback_data="back_main")]
        ]
        creply_markup = InlineKeyboardMarkup(ckeyboard)
        # checks if user is in enter phone number step
        if self.user_status[chat_id] == "enter_phone_number":
            # validates whether the pnr is valid or not
            valid = False
            if(len(update.message.text) == 10):
                if(update.message.text[0:1] == "0"):
                    valid = True
            # if the pnr is valid then the bot will ask for a code
            if valid:
                ctx.bot.send_message(
                    chat_id, messages["enter_operator_code"], reply_markup=creply_markup)
                self.user_status[chat_id] = "enter_operator_code"
                self.user_inputs[chat_id] = [update.message.text]
            # if the phone is invalid then the bot will ask again
            else:
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])

        # checks if user is in enter operator code
        elif self.user_status[chat_id] == "enter_operator_code" and len(self.user_inputs[chat_id]) > 0:
            # validates whether the pnr is valid or not
            valid = False
            if(len(update.message.text) == 4 and update.message.text.isdigit()):
                valid = True
            # if the code is valid then the bot will try to authenticate the operator
            if valid:
                creds = self.operator_authenticator(
                    self.user_inputs[chat_id][0], update.message.text)
                # if auth passes
                if(len(creds) == 4):
                    self.user_inputs[chat_id] = creds
                    keyboard = []

                    for line in open(f"{CURRENT_DIR}/data/locations.txt", "rb"):
                        keyboard.append([line.decode().rstrip()])
                    keyboard.append([messages["back"]])
                    reply_markup = ReplyKeyboardMarkup(
                        keyboard, resize_keyboard=True, one_time_keyboard=True)
                    ctx.bot.send_message(
                        chat_id, messages["welcome_operator"]+" "+self.user_inputs[chat_id][0])
                    ctx.bot.send_message(chat_id, messages["select_location"],
                                         reply_markup=reply_markup)

                    self.user_status[chat_id] = "select_source"
                else:
                    ctx.bot.send_message(
                        chat_id, messages["wrong_credentials"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
                    self.user_status.pop(chat_id, None)
                    self.user_section.pop(chat_id, None)
                    self.user_inputs.pop(chat_id, None)
                    self.operator_menu(update, ctx)

            else:
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])
                self.user_status.pop(chat_id, None)
                self.user_section.pop(chat_id, None)
                self.user_inputs.pop(chat_id, None)
                self.operator_menu(update, ctx)

        # checks if user is in select source step
        elif self.user_status[chat_id] == "select_source":
            locations = []
            with open(f"{CURRENT_DIR}/data/locations.txt", "rb") as f:
                for line in f:
                    locations.append([line.decode().rstrip()])
                locations.append([messages["back"]])
                f.close()

                # validates whether the source is valid or not
                valid = False
                for line in open(f"{CURRENT_DIR}/data/locations.txt", "rb"):
                    if line.decode().rstrip() == update.message.text:
                        self.user_inputs[chat_id].append(update.message.text)
                        valid = True
                        break

                # if the source is valid then the bot will ask for destination
                if valid:
                    ctx.bot.send_message(chat_id, messages["select_destination"],
                                         reply_markup=ReplyKeyboardMarkup(locations, one_time_keyboard=True))

                    self.user_status[chat_id] = "select_destination"

                # if the source is invalid then the bot will ask again
                else:
                    ctx.bot.send_message(chat_id, messages["wrong_location"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
                    ctx.bot.send_message(chat_id, messages["select_location"],
                                         reply_markup=ReplyKeyboardMarkup(locations, one_time_keyboard=True))

        # checks if user is in select destination step
        elif self.user_status[chat_id] == "select_destination" and len(self.user_inputs[chat_id]) > 4:
            # validates whether the destination is valid or not
            valid = False
            for line in open(f"{CURRENT_DIR}/data/locations.txt", "rb"):
                if line.decode().rstrip() == update.message.text:
                    self.user_inputs[chat_id].append(update.message.text)
                    valid = True
                    break

            # if the destination is valid then the bot will ask for trip date
            if valid:
                ctx.bot.send_message(
                    chat_id, messages["select_trip_date"], reply_markup=creply_markup)
                self.user_status[chat_id] = "select_trip_date"

            # if the destination is invalid then the bot will ask again
            else:
                ctx.bot.send_message(chat_id, messages["wrong_location"])
                ctx.bot.send_message(chat_id, messages["try_again"])
                ctx.bot.send_message(chat_id, messages["select_destination"],
                                     reply_markup=ReplyKeyboardMarkup(locations, one_time_keyboard=True))

        # checks if user is in select trip date step
        elif self.user_status[chat_id] == "select_trip_date" and len(self.user_inputs[chat_id]) > 5:
            try:
                # validates whether the date is valid or not
                if self.date_validator(update.message.text):
                    # if the date is valid then the bot will ask for passenger number
                    ctx.bot.send_message(
                        chat_id, messages["enter_number_of_passengers"], reply_markup=creply_markup)
                    self.user_status[chat_id] = "enter_number_of_passengers"
                    self.user_inputs[chat_id].append(update.message.text)

                # if the date is invalid then the bot will ask again
                else:
                    ctx.bot.send_message(chat_id, messages["wrong_date"])
                    ctx.bot.send_message(chat_id, messages["try_again"])

            except Exception as e:
                print(e)
                # if the date has an error then the bot will redirect user back to operator_menu
                ctx.bot.send_message(chat_id, messages["wrong_date"])
                ctx.bot.send_message(chat_id, messages["try_again"])
                self.user_status.pop(chat_id, None)
                self.user_section.pop(chat_id, None)
                self.user_inputs.pop(chat_id, None)
                self.operator_menu(update, ctx)

        # checks if user is in enter number of passengers step
        elif self.user_status[chat_id] == "enter_number_of_passengers" and len(self.user_inputs[chat_id]) > 6:
            try:
                # validates whether the number is valid or not
                valid = False
                if(update.message.text.isdigit()):
                    valid = True

                   # if the date is valid then the bot wil ask for confirmation
                if valid:
                    self.user_inputs[chat_id].append(update.message.text)
                    reply_markup = ReplyKeyboardMarkup(
                        [[messages["cancel"], messages["submit"]]], resize_keyboard=True, one_time_keyboard=True)
                    ctx.bot.send_message(chat_id, messages["confirm"])
                    print(self.user_inputs[chat_id])

                    ctx.bot.send_message(chat_id, messages["name"]
                                         + ": "+self.user_inputs[chat_id][0]
                                         + "\n"+messages["bus"]+": " +
                                         self.user_inputs[chat_id][3]
                                         + "\n"+messages["source"]+": " +
                                         self.user_inputs[chat_id][4]
                                         + "\n"+messages["destination"] +
                                         ": "+self.user_inputs[chat_id][5]
                                         + "\n"+messages["date"]+": " +
                                         self.user_inputs[chat_id][6]
                                         + "\n"+messages["seats"]+": " +
                                         self.user_inputs[chat_id][7],
                                         reply_markup=reply_markup)
                    self.user_status[chat_id] = "selling_confirmation"
                else:
                    ctx.bot.send_message(chat_id, messages["wrong_number"])
                    ctx.bot.send_message(chat_id, messages["try_again"])

            except Exception as e:
                print(e)
                # if the number has an error then the bot will redirect user back to operator menu

                ctx.bot.send_message(chat_id, messages["wrong_number"])
                ctx.bot.send_message(chat_id, messages["try_again"])
                self.user_status.pop(chat_id, None)
                self.user_section.pop(chat_id, None)
                self.user_inputs.pop(chat_id, None)
                self.operator_menu(update, ctx)

        # checks if user is in confirmation
        elif self.user_status[chat_id] == "selling_confirmation" and len(self.user_inputs[chat_id]) > 7:
            try:
                # validates whether the confirmation is submitted or not
                if(update.message.text == messages["submit"]):
                    if(self.add_seat_to_excel(self.user_inputs[chat_id])):
                        keyboard = [
                            [InlineKeyboardButton(
                                messages["back_to_main"], callback_data="back_main")]
                        ]
                        reply_markup = InlineKeyboardMarkup(keyboard)
                        ctx.bot.send_message(
                            chat_id, messages["seat_submitted"], reply_markup=reply_markup)
                        dname = self.user_inputs[chat_id][0]
                        dphone = self.user_inputs[chat_id][1]
                        ddate = self.user_inputs[chat_id][6]
                        dsource = self.user_inputs[chat_id][4]
                        ddestination = self.user_inputs[chat_id][5]
                        dseats = self.user_inputs[chat_id][7]
                        dbus = self.user_inputs[chat_id][3]
                        for adm in ADMIN:
                            files = {"document": open(
                                f"data/seats.xlsx", 'rb')}
                            resp = requests.post(
                                "https://api.telegram.org/bot" + TOKEN+"/sendDocument?chat_id="+str(adm), files=files)
                            print(resp)
                            ctx.bot.send_message(adm, messages["new_operator_selling_seats"]
                                                 #  + "\n@" +
                                                 #  update.message.from_user["username"]+": "
                                                 + "\n"+messages["name"]+": " +
                                                 dname
                                                 + "\n" +
                                                 messages["phone"]+": 0" +
                                                 str(int(dphone))
                                                 + "\n"+messages["bus"]+": " +
                                                 dbus
                                                 + "\n"+messages["source"]+": " +
                                                 dsource
                                                 + "\n"+messages["destination"] +
                                                 ": "+ddestination
                                                 + "\n"+messages["date"]+": " +
                                                 ddate
                                                 + "\n"+messages["seats"]+": "+dseats)
                        self.user_status.pop(chat_id, None)
                        self.user_section.pop(chat_id, None)
                        self.user_inputs.pop(chat_id, None)

                    else:
                        ctx.bot.send_message(chat_id, messages["wrong_value"])
                        ctx.bot.send_message(chat_id, messages["try_again"])
                        self.user_status.pop(chat_id, None)
                        self.user_section.pop(chat_id, None)
                        self.user_inputs.pop(chat_id, None)
                        self.operator_menu(update, ctx)
                else:
                    self.user_status.pop(chat_id, None)
                    self.user_section.pop(chat_id, None)
                    self.user_inputs.pop(chat_id, None)
                    self.operator_menu(update, ctx)
            except Exception as e:
                print("exception")
                if str(e) == "Chat not found":
                    print("Chat not fouond")
                elif str(e) == "Forbidden: bot was blocked by the user":
                    print("Bot Blocked by user")
                else:
                    print(e)
                    # if the date has an error then the bot will redirect user back to passenger menu
                    #ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])

    # handles the process of selling seats
    def upload_trip_process(self, update, ctx, chat_id, messages):
        ckeyboard = [
            [InlineKeyboardButton(messages["back_to_main"],
                                  callback_data="back_main")]
        ]
        creply_markup = InlineKeyboardMarkup(ckeyboard)
        # checks if user is in enter phone number step
        if self.user_status[chat_id] == "enter_phone_number":
            # validates whether the pnr is valid or not
            valid = False
            if(len(update.message.text) == 10):
                if(update.message.text[0:1] == "0"):
                    valid = True
            # if the pnr is valid then the bot will ask for a code
            if valid:
                ctx.bot.send_message(
                    chat_id, messages["enter_operator_code"], reply_markup=creply_markup)
                self.user_status[chat_id] = "enter_operator_code"
                self.user_inputs[chat_id] = [update.message.text]
            # if the phone is invalid then the bot will ask again
            else:
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])

        # checks if user is in enter operator code
        elif self.user_status[chat_id] == "enter_operator_code" and len(self.user_inputs[chat_id]) > 0:
            # validates whether the pnr is valid or not
            valid = False
            if(len(update.message.text) == 4 and update.message.text.isdigit()):
                valid = True
            # if the code is valid then the bot will try to authenticate the operator
            if valid:
                creds = self.operator_authenticator(
                    self.user_inputs[chat_id][0], update.message.text)
                # if auth passes
                if(len(creds) == 4):
                    self.user_inputs[chat_id] = creds

                    ctx.bot.send_message(
                        chat_id, messages["welcome_operator"]+" "+self.user_inputs[chat_id][0])
                    ctx.bot.send_message(
                        chat_id, messages["weekly_trip_message"], reply_markup=creply_markup)
                    self.user_status[chat_id] = "weekly_trip_message"
                else:
                    ctx.bot.send_message(
                        chat_id, messages["wrong_credentials"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
                    self.user_status.pop(chat_id, None)
                    self.user_section.pop(chat_id, None)
                    self.user_inputs.pop(chat_id, None)
                    self.operator_menu(update, ctx)

            else:
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])
                self.user_status.pop(chat_id, None)
                self.user_section.pop(chat_id, None)
                self.user_inputs.pop(chat_id, None)
                self.operator_menu(update, ctx)

        # checks if user is in enter message step
        elif self.user_status[chat_id] == "weekly_trip_message" and len(self.user_inputs[chat_id]) > 3:
            # validates whether the message is valid or not
            valid = False
            if(len(update.message.text) <= 500):
                valid = True
            # if the message is valid then the bot will ask for a message
            if valid:
                ctx.bot.send_message(
                    chat_id, messages["upload_trip_document"], reply_markup=creply_markup)
                self.user_status[chat_id] = "upload_weekly_file"
                self.user_inputs[chat_id].append(update.message.text)
                print(self.user_inputs[chat_id])
            # if the message is invalid then the bot will ask again
            else:
                ctx.bot.send_message(chat_id, messages["long_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])

        # checks if user is in select message step
        elif self.user_status[chat_id] == "weekly_trip_confirmation" and len(self.user_inputs[chat_id]) > 5:
            try:
                # validates whether the confirmation is submitted or not
                if(update.message.text == messages["submit"]):
                    # if the user submits the confirmation then the system will send the data to the admin and delete the picture
                    filename = self.user_inputs[chat_id][5].file_path.split(
                        "/")[-1]
                    self.user_inputs[chat_id][5].download("data/"+filename)
                    print(self.user_inputs[chat_id])
                    dname = self.user_inputs[chat_id][0]
                    dphone = self.user_inputs[chat_id][1]
                    dmessage = self.user_inputs[chat_id][4]
                    ctx.bot.send_message(
                        chat_id, messages["weekly_trip_submitted"], reply_markup=creply_markup)
                    for adm in ADMIN:
                        files = {"document": open(f"data/{filename}", "rb")}
                        ctx.bot.send_message(adm,  "#new_weekly_trip_submitted"+"\n"+messages["name"]+": " +
                                             dname + "\n" +
                                             messages["phone"]+": 0" +
                                             str(int(dphone))+"\n" +
                                             messages["sent_message"]+": "+dmessage)

                        resp = requests.post("https://api.telegram.org/bot" +
                                             TOKEN+"/sendDocument?chat_id="+str(adm), files=files)
                    print(resp.status_code)
                    self.user_status.pop(chat_id, None)
                    self.user_section.pop(chat_id, None)
                    self.user_inputs.pop(chat_id, None)
                    os.remove(f"{CURRENT_DIR}/data/{filename}")

                else:
                    self.user_status.pop(chat_id, None)
                    self.user_section.pop(chat_id, None)
                    self.user_inputs.pop(chat_id, None)
                    self.operator_menu(update, ctx)
                valid = False
            except Exception as e:
                print("exception")
                if str(e) == "Chat not found":
                    print("Chat not fouond")
                elif str(e) == "Forbidden: bot was blocked by the user":
                    print("Bot Blocked by user")
                else:
                    print(e)
                    # if the date has an error then the bot will redirect user back to passenger menu
                    #ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])

        elif self.user_status[chat_id] == "upload_weekly_file":
            # if the user sends message instead of image, the system will ask again
            ctx.bot.send_message(
                chat_id, messages["send_file"], reply_markup=creply_markup)

       # handles the process of verifying passengers
    def verify_passenger_process(self, update, ctx, chat_id, messages):
        ckeyboard = [
            [InlineKeyboardButton(messages["back_to_main"],
                                  callback_data="back_main")]
        ]
        creply_markup = InlineKeyboardMarkup(ckeyboard)
        # checks if user is in enter phone number step
        if self.user_status[chat_id] == "enter_phone_number":
            # validates whether the pnr is valid or not
            valid = False
            if(len(update.message.text) == 10):
                if(update.message.text[0:1] == "0"):
                    valid = True
            # if the pnr is valid then the bot will ask for a code
            if valid:
                ctx.bot.send_message(
                    chat_id, messages["enter_operator_code"], reply_markup=creply_markup)
                self.user_status[chat_id] = "enter_operator_code"
                self.user_inputs[chat_id] = [update.message.text]
            # if the phone is invalid then the bot will ask again
            else:
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])

        # checks if user is in enter operator code
        elif self.user_status[chat_id] == "enter_operator_code" and len(self.user_inputs[chat_id]) > 0:
            # validates whether the pnr is valid or not
            valid = False
            if(len(update.message.text) == 4 and update.message.text.isdigit()):
                valid = True
            # if the code is valid then the bot will try to authenticate the operator
            if valid:
                creds = self.operator_authenticator(
                    self.user_inputs[chat_id][0], update.message.text)
                # if auth passes
                if(len(creds) == 4):
                    self.user_inputs[chat_id] = creds

                    ctx.bot.send_message(
                        chat_id, messages["welcome_operator"]+" "+self.user_inputs[chat_id][0])
                    chat_id = update.effective_chat.id
                    ctx.bot.send_message(
                        chat_id, messages["enter_customer_pnr_no"], reply_markup=creply_markup)

                    self.user_status[chat_id] = "enter_pnr"
                    self.user_section[chat_id] = "verify_passenger"

                else:
                    ctx.bot.send_message(
                        chat_id, messages["wrong_credentials"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
                    self.user_status.pop(chat_id, None)
                    self.user_section.pop(chat_id, None)
                    self.user_inputs.pop(chat_id, None)
                    self.operator_menu(update, ctx)

            else:
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])
                self.user_status.pop(chat_id, None)
                self.user_section.pop(chat_id, None)
                self.user_inputs.pop(chat_id, None)
                self.operator_menu(update, ctx)

        elif self.user_status[chat_id] == "enter_pnr" and len(self.user_inputs[chat_id]) > 3:
            # validates whether the pnr is valid or not
            valid = False
            if(len(update.message.text) == 8):
                valid = True
            # if the pnr is valid then the bot will ask for a message
            if valid:
                ctx.bot.send_message(
                    chat_id, messages["enter_passenger_phone"], reply_markup=creply_markup)
                self.user_status[chat_id] = "enter_phone"
                self.user_inputs[chat_id].append(update.message.text)
            # if the pnr is invalid then the bot will redirect user back to passenger menu
            else:
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])
                self.user_status.pop(chat_id, None)
                self.user_section.pop(chat_id, None)
                self.user_inputs.pop(chat_id, None)
                self.passenger_menu(update, ctx)

        # checks if user is in enter name step
        elif self.user_status[chat_id] == "enter_phone" and len(self.user_inputs[chat_id]) > 4:
            try:
                valid = False
                if(len(update.message.text) == 10):
                    if(update.message.text[0:1] == "0" and update.message.text.isdigit()):
                        valid = True
                if valid:
                    cds = self.get_info_by_pnr(
                        update.message.text, self.user_inputs[chat_id][4])
                    if len(cds) != 0:
                        ctx.bot.send_message(
                            chat_id, messages["trip_is_verified"])
                        ctx.bot.send_message(chat_id,
                                             messages["name"] + " : " + cds[1]
                                             + "\n" +
                                             messages["date"] + " : " + cds[0]
                                             + "\n" +
                                             messages["source"] +
                                             " : " + cds[2]
                                             + "\n" +
                                             messages["destination"] +
                                             " : " + cds[3]
                                             + "\n"+messages["bus"] + " : " + cds[4], reply_markup=creply_markup)
                        self.user_status.pop(chat_id, None)
                        self.user_section.pop(chat_id, None)
                        self.user_inputs.pop(chat_id, None)

                    else:
                        ctx.bot.send_message(
                            chat_id, messages["wrong_pnr_credentials"])
                        ctx.bot.send_message(chat_id, messages["try_again"])
                        self.user_status.pop(chat_id, None)
                        self.user_section.pop(chat_id, None)
                        self.user_inputs.pop(chat_id, None)
                        self.operator_menu(update, ctx)
                else:
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
            except Exception as e:
                print(e)
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])

    def add_pnr_process(self, update, ctx, chat_id, messages):
        ckeyboard = [
            [InlineKeyboardButton(messages["back_to_main"],
                                  callback_data="back_main")]
        ]
        creply_markup = InlineKeyboardMarkup(ckeyboard)
        # checks if user is in enter phone number step
        if self.user_status[chat_id] == "enter_phone_number":
            # validates whether the pnr is valid or not
            valid = False
            if(len(update.message.text) == 10):
                if(update.message.text[0:1] == "0"):
                    valid = True
            # if the pnr is valid then the bot will ask for a code
            if valid:
                ctx.bot.send_message(
                    chat_id, "Enter the booking ID", reply_markup=creply_markup)
                self.user_status[chat_id] = "enter_booking_id"
                self.user_inputs[chat_id] = [update.message.text]
            # if the phone is invalid then the bot will ask again
            else:
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])

        # checks if user is in enter operator code
        elif self.user_status[chat_id] == "enter_booking_id" and len(self.user_inputs[chat_id]) > 0:
            # validates whether the pnr is valid or not
            valid = False
            if(len(update.message.text) == 8 and update.message.text.isdigit()):
                valid = True
            # if the code is valid then the bot will try to authenticate the operator
            if valid:

                self.user_inputs[chat_id].append(update.message.text)

                ctx.bot.send_message(
                    chat_id, "Enter the PNR that you want to assign", reply_markup=creply_markup)

                self.user_status[chat_id] = "enter_pnr"

            else:
                ctx.bot.send_message(chat_id, messages["wrong_value"])
                ctx.bot.send_message(chat_id, messages["try_again"])

        elif self.user_status[chat_id] == "enter_pnr" and len(self.user_inputs[chat_id]) > 1:
            try:
                # validates whether the pnr is valid or not
                valid = False
                if(len(update.message.text) == 8):
                    valid = True
                # if the pnr is valid then the bot will ask for a message
                if valid:
                    if self.pnr_assigner(self.user_inputs[chat_id][0], self.user_inputs[chat_id][1], update.message.text):
                        ctx.bot.send_message(
                            chat_id, "Successfully Updated PNR")
                        for adm in ADMIN:
                            files = {"document": open(
                                f"data/bookings.xlsx", 'rb')}
                            resp = requests.post(
                                "https://api.telegram.org/bot"+TOKEN+"/sendDocument?chat_id="+str(adm), files=files)
                            print(resp)

                        self.user_status.pop(chat_id, None)
                        self.user_section.pop(chat_id, None)
                        self.user_inputs.pop(chat_id, None)
                        self.admin_menu(update, ctx)
                    else:
                        ctx.bot.send_message(
                            chat_id, "PNR not Updated")
                        self.user_status.pop(chat_id, None)
                        self.user_section.pop(chat_id, None)
                        self.user_inputs.pop(chat_id, None)
                        self.admin_menu(update, ctx)

                # if the pnr is invalid then the bot will redirect user back to passenger menu
                else:
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])
            except Exception as e:
                print("exception")
                if str(e) == "Chat not found":
                    print("Chat not fouond")
                elif str(e) == "Forbidden: bot was blocked by the user":
                    print("Bot Blocked by user")
                else:
                    print(e)
                    # if the date has an error then the bot will redirect user back to passenger menu
                    #ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["wrong_value"])
                    ctx.bot.send_message(chat_id, messages["try_again"])

    def upload_proof(self, update, ctx):
        chat_id = update.effective_chat.id

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages
        ckeyboard = [[InlineKeyboardButton(
            messages["back_to_main"], callback_data="back_main")]]
        creply_markup = InlineKeyboardMarkup(ckeyboard)

        ctx.bot.send_message(
            chat_id, messages["enter_booking_pnr_no"], reply_markup=creply_markup)

        self.user_status[chat_id] = "enter_pnr"
        self.user_section[chat_id] = "payment_proof"

    def book_bus(self, update, ctx):
        chat_id = update.effective_chat.id

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages

        keyboard = []

        for line in open(f"{CURRENT_DIR}/data/locations.txt", "rb"):
            keyboard.append([line.decode().rstrip()])
        keyboard.append([messages["back"]])

        reply_markup = ReplyKeyboardMarkup(
            keyboard, resize_keyboard=True, one_time_keyboard=True)

        ctx.bot.send_message(chat_id, messages["select_location"],
                             reply_markup=reply_markup)

        self.user_status[chat_id] = "select_source"
        self.user_section[chat_id] = "book_bus"

    def sell_seats(self, update, ctx):
        chat_id = update.effective_chat.id
        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages
        ckeyboard = [[InlineKeyboardButton(
            messages["back_to_main"], callback_data="back_main")]]
        creply_markup = InlineKeyboardMarkup(ckeyboard)
        ctx.bot.send_message(
            chat_id, messages["enter_phone_number"], reply_markup=creply_markup)
        self.user_status[chat_id] = "enter_phone_number"
        self.user_section[chat_id] = "sell_seats"

    def upload_weekly_trip(self, update, ctx):
        chat_id = update.effective_chat.id

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages
        ckeyboard = [[InlineKeyboardButton(
            messages["back_to_main"], callback_data="back_main")]]
        creply_markup = InlineKeyboardMarkup(ckeyboard)
        ctx.bot.send_message(
            chat_id, messages["enter_phone_number"], reply_markup=creply_markup)
        self.user_status[chat_id] = "enter_phone_number"
        self.user_section[chat_id] = "upload_trip"

    def add_pnr(self, update, ctx):
        chat_id = update.effective_chat.id

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages
        ckeyboard = [[InlineKeyboardButton(
            messages["back_to_main"], callback_data="back_main")]]
        creply_markup = InlineKeyboardMarkup(ckeyboard)
        ctx.bot.send_message(
            chat_id, "Enter the phone number of the customer", reply_markup=creply_markup)
        self.user_status[chat_id] = "enter_phone_number"
        self.user_section[chat_id] = "add_pnr"

    def verify_passenger(self, update, ctx):
        chat_id = update.effective_chat.id

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages
        ckeyboard = [[InlineKeyboardButton(
            messages["back_to_main"], callback_data="back_main")]]
        creply_markup = InlineKeyboardMarkup(ckeyboard)
        ctx.bot.send_message(
            chat_id, messages["enter_phone_number"], reply_markup=creply_markup)
        self.user_status[chat_id] = "enter_phone_number"
        self.user_section[chat_id] = "verify_passenger"

    def operator_menu(self, update, ctx):
        chat_id = update.effective_chat.id

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages

        reply_markup = ReplyKeyboardMarkup(
            [[messages["sell_seats"]], [messages["upload_list"]],
             [messages["verify_passenger"]], [messages["back"]]], resize_keyboard=True, one_time_keyboard=True)

        ctx.bot.send_message(chat_id, messages["operator_choice"],
                             reply_markup=reply_markup)

    def passenger_menu(self, update, ctx):
        chat_id = update.effective_chat.id

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages

        reply_markup = ReplyKeyboardMarkup(
            [[messages["book_bus"]], [messages["proof"]], [messages["back"]]], resize_keyboard=True, one_time_keyboard=True)

        ctx.bot.send_message(
            chat_id, messages["passenger_choice"], reply_markup=reply_markup)

    def refer(self, update, ctx):
        chat_id = update.effective_chat.id
        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages
        #ctx.bot.send_message(chat_id, messages["referal_link"]+ " www.google.com")
        keyboard = [
            [InlineKeyboardButton(messages["back"], callback_data="back_main")]
        ]

        reply_markup = InlineKeyboardMarkup(keyboard)

        if update.message:
            update.message.reply_text(messages["actual_referal_link"],
                                      reply_markup=reply_markup)
        else:
            update.callback_query.edit_message_text(update.callback_query.message.text,
                                                    reply_markup=reply_markup)

    def about(self, update, ctx):
        chat_id = update.effective_chat.id
        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages
        #ctx.bot.send_message(chat_id, messages["referal_link"]+ " www.google.com")
        keyboard = [
            [InlineKeyboardButton(messages["back"], callback_data="back_main")]
        ]

        reply_markup = InlineKeyboardMarkup(keyboard)

        if update.message:
            update.message.reply_text(messages["about_guzo"],
                                      reply_markup=reply_markup)
        else:
            update.callback_query.edit_message_text(update.callback_query.message.text,
                                                    reply_markup=reply_markup)

    def rules(self, update, ctx):
        chat_id = update.effective_chat.id
        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages
        #ctx.bot.send_message(chat_id, messages["referal_link"]+ " www.google.com")
        keyboard = [
            [InlineKeyboardButton(messages["back"], callback_data="back_main")]
        ]

        reply_markup = InlineKeyboardMarkup(keyboard)

        if update.message:
            update.message.reply_text(messages["guzo_rules"],
                                      reply_markup=reply_markup)
        else:
            update.callback_query.edit_message_text(update.callback_query.message.text,
                                                    reply_markup=reply_markup)

    def contactus(self, update, ctx):
        chat_id = update.effective_chat.id
        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages
        #ctx.bot.send_message(chat_id, messages["referal_link"]+ " www.google.com")
        keyboard = [
            [InlineKeyboardButton(messages["back"], callback_data="back_main")]
        ]

        reply_markup = InlineKeyboardMarkup(keyboard)

        if update.message:
            update.message.reply_text(messages["guzo_contacts"],
                                      reply_markup=reply_markup)
        else:
            update.callback_query.edit_message_text(update.callback_query.message.text,
                                                    reply_markup=reply_markup)

    def start(self, update, ctx):
        user = update.message
        chat_id = update.effective_chat.id

        if not update.message:
            user = update.callback_query

        if chat_id not in self.messages:
            self.messages[chat_id] = "en"

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages

        reply_markup = ReplyKeyboardMarkup(
            [["English 🇬🇧", "አማርኛ 🇪🇹"]], resize_keyboard=True)

        # ctx.bot.send_message(chat_id, text = "Welcome to Guzo Bot, እንኳን ወደ ጉዞ ቦት መጣችሁ\nChoose Language to Proceed. ለመቀጠል ቋንቋ ይምረጡ።")
        self.user_status.pop(chat_id, None)
        self.user_inputs.pop(chat_id, None)
        self.user_section.pop(chat_id, None)
        files = {"photo": open("data/images/guzo_welcome.jpg", 'rb')}
        resp = requests.post("https://api.telegram.org/bot" +
                             TOKEN+"/sendPhoto?chat_id="+str(chat_id), files=files)
        print(resp.status_code)
        update.message.reply_text(
            "🚌  Welcome to Guzo Bus Ethiopia Bot\nእንኳን ወደ ጉዞ ባስ ኢትዮጵያ ቦት በሰላም መጡ\n\n💺Choose Language to Proceed\nለመቀጠል ቋንቋ ይምረጡ:-", reply_markup=reply_markup)

        #reply_markup = InlineKeyboardMarkup(keyboard)

    def error(self, update, ctx):
        self.start(update, ctx)
        logger.warning('Update "%s" caused error "%s"', update, ctx.error)

    def admin_menu(self, update, ctx):
        user = update.message
        chat_id = update.effective_chat.id
        if not update.message:
            user = update.callback_query
        reply_markup = ReplyKeyboardMarkup(
            [["Add PNR"], ["Seats", "Bookings"], ["Operators", "Locations"], ["English", "Amharic"]], resize_keyboard=True)

        ctx.bot.send_message(chat_id, text="Welcome Admin",
                             reply_markup=reply_markup)
        self.user_status.pop(chat_id, None)
        self.user_inputs.pop(chat_id, None)
        self.user_section.pop(chat_id, None)

    def main_menu(self, update, ctx):
        user = update.message
        chat_id = update.effective_chat.id

        if not update.message:
            user = update.callback_query

        if chat_id not in self.messages:
            self.messages[chat_id] = "en"

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages

        self.user_status.pop(chat_id, None)
        self.user_section.pop(chat_id, None)
        self.user_inputs.pop(chat_id, None)
        reply_markup = ReplyKeyboardMarkup(
            [[messages["quest_passenger"]], [messages["quest_operator"]], [messages["referal_link"], messages["about_us"]], [messages["rules"], messages["contact_us"]], [messages["language_switcher"]]], resize_keyboard=True, one_time_keyboard=True
        )

        if update.message:
            update.message.reply_text(messages["welcome_message"] % user.chat.first_name,
                                      reply_markup=reply_markup)
        else:
            ctx.bot.send_message(
                chat_id, messages["clean_welcome_message"], reply_markup=reply_markup)

    def date_validator(self, date):
        valid = False
        if(len(date) == 10):
            if(date[2:3] == "/" and date[5:6] == "/"):
                if(date[0:2].isdigit()):
                    if(int(date[0:2]) > 0 and int(date[0:2]) <= 13):
                        if(date[3:5].isdigit()):
                            if(int(date[3:5]) > 0 and int(date[3:5]) <= 30):
                                if(date[6:].isdigit()):
                                    if(int(date[6:]) > 2013 and int(date[6:]) <= 2015):
                                        valid = True
        return valid

    def date_checker(self, tempdate, inputdate):
        if(inputdate[3:5] == tempdate[8:10] and inputdate[0:2] == tempdate[5:7] and inputdate[6:] == tempdate[0:4]):
            return True
        else:
            return False

    # the main function that return the number of available buses
    def seat_match_maker(self, date, source, destination, seats):
        abuses = []
        try:
            print("destination: "+destination)
            excel_file = 'data/seats.xlsx'
            df = pd.read_excel(excel_file)
            res = df.where(df['destination'] == destination).dropna()
            # print(res)
            nbuses = 0
            for (xdate, xsource, xdestination, xseats, xbus, xprice) in res.values:
                #print(str(xdate)+" - "+xsource+" - "+xdestination+" - "+ str(xseats)+ " - "+xbus)
                print("source: "+str(xsource == source))
                print(xsource+" : "+source)
                print("--")
                print("destination: "+str(xdestination == destination))
                print(xdestination+" : "+destination)
                print("--")
                print("seats: "+str(xseats >= int(seats)))

                if(xsource == source):
                    # tempdate = str(datetime.strptime(
                    #    str(xdate), '%Y-%m-%d %H:%M:%S'))
                    # if(self.date_checker(tempdate, date) == True):
                    # print(str(xdate))
                    if(xseats >= int(seats)):
                        if xbus not in abuses and len(abuses) < 3:
                            if(xprice > 0):
                                print(xbus+" Appended")
                                abuses.append(xbus)
                                nbuses += 1
            return abuses
        except Exception as e:
            print(e)
            return abuses

    # decrement seats based on user input
    def seat_reserver(self, date, source, destination, seats, bus):
        busprice = 0
        try:
            excel_file = 'data/seats.xlsx'
            df = pd.read_excel(excel_file)
            res = df.where(df['seats'] >= int(seats)).dropna()

            chosenseats = -1
            chosenid = -1
            for i in range(len(res.index)):
                #xdate = res.values[i][0]
                xsource = res.values[i][1]
                xdestination = res.values[i][2]
                xseats = res.values[i][3]
                xbus = res.values[i][4]
                xbusprice = res.values[i][5]
                print(xsource+" - "+str(xseats) +
                      " - "+xbus+" - "+str(xbusprice))
                print("temp id - " + str(res.index.values[i]))
                if(xsource == source and xdestination == destination):
                    # tempdate = str(datetime.strptime(
                    #   str(xdate), '%Y-%m-%d'))
                    # if(self.date_checker(tempdate, date) == True):
                    if(xbus == bus and xbusprice > 0):
                        chosenseats = xseats
                        chosenid = res.index.values[i]
                        busprice = xbusprice
                        break

            # load excel file
            workbook = load_workbook(filename="data/seats.xlsx")

            # open workbook
            sheet = workbook.active

            # modify the desired cell
            print("chosen id - "+str(chosenid))
            print(chosenseats)
            print(busprice)
            if chosenid != -1 and chosenseats != -1:
                sheet["D"+str(chosenid+2)] = chosenseats - int(seats)
                # save the file
                workbook.save(filename="data/seats.xlsx")
                return busprice

            return busprice

        except Exception as e:
            print(e)
            return busprice

    def pnr_assigner(self, phone, bookingid, pnr):
        try:
            excel_file = 'data/bookings.xlsx'
            df = pd.read_excel(excel_file)
            res = df.where(df['no'] == int(bookingid)).dropna()
            chosenid = -1
            chosenpnr = "-"
            for i in range(len(res.index)):
                #xdate = res.values[i][0]
                tel = res.values[i][3]
                print("tel - " + str(tel))
                if(tel == int(phone)):
                    chosenpnr = pnr
                    chosenid = res.index.values[i]

            # load excel file
            workbook = load_workbook(filename="data/bookings.xlsx")

            # open workbook
            sheet = workbook.active

            # modify the desired cell
            print("chosen pnr - "+str(chosenpnr))
            if chosenpnr != "-" and chosenid != -1:
                sheet["L"+str(chosenid+2)] = chosenpnr
                # save the file
                workbook.save(filename="data/bookings.xlsx")
                return True
            return False
        except Exception as e:
            print(e)
            return False

    def add_bookings_to_excel(self, info_list):
        try:
            print(info_list)
            print(["no", "bookingdate", "fullname", "tel", "sourcecity", "destinationcity",
                   "traveldate", "seats", "bus", "price", "totalprice", "bookngpnr"])
            date = info_list[0]
            now = datetime.now()
            currentdate = now.strftime('%m/%d/%Y')
            fdate = str(int(date[0:2]))+"/" + \
                str(int(date[3:5]))+"/"+str(int(date[6:]))
            cfdate = str(int(currentdate[0:2]))+"/" + \
                str(int(currentdate[3:5]))+"/"+str(int(currentdate[6:]))
            df = pd.DataFrame([[int(info_list[10]), cfdate, info_list[1], int(info_list[2]), info_list[3], info_list[4], fdate, int(info_list[6]), info_list[7], int(info_list[8]), int(
                info_list[9]), -1]], columns=["no", "bookingdate", "fullname", "tel", "sourcecity", "destinationcity", "traveldate", "seats", "bus", "price", "totalprice", "bookngpnr"])
            book = load_workbook(f'{CURRENT_DIR}/data/bookings.xlsx')
            writer = pd.ExcelWriter(
                f'{CURRENT_DIR}/data/bookings.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(
                writer, startrow=writer.sheets['Sheet1'].max_row, index=False, header=False)
            writer.save()
            return True
        except Exception as e:
            print(e)
            return False

    def add_seat_to_excel(self, info_list):
        tempdate = info_list[6]
        date = str(int(tempdate[0:2]))+"/" + \
            str(int(tempdate[3:5]))+"/"+str(int(tempdate[6:]))
        source = info_list[4]
        destination = info_list[5]
        seats = int(info_list[7])
        bus = info_list[3]
        busprice = 0
        try:
            excel_file = 'data/seats.xlsx'
            df = pd.read_excel(excel_file)
            res = df.where(df['bus'] == bus).dropna()

            chosenseats = -1
            chosenid = -1
            for i in range(len(res.index)):
                #xdate = res.values[i][0]
                xsource = res.values[i][1]
                xdestination = res.values[i][2]
                xseats = res.values[i][3]
                xbus = res.values[i][4]
                xbusprice = res.values[i][5]
                print(xsource+" - "+str(xseats) +
                      " - "+xbus+" - "+str(xbusprice))
                print("temp id - " + str(res.index.values[i]))
                if(xsource == source and xdestination == destination):
                    # tempdate = str(datetime.strptime(
                    #   str(xdate), '%Y-%m-%d'))
                    # if(self.date_checker(tempdate, date) == True):
                    if(xbus == bus and xbusprice > 0):
                        chosenseats = xseats
                        chosenid = res.index.values[i]
                        busprice = xbusprice
                        break

            # load excel file
            workbook = load_workbook(filename="data/seats.xlsx")

            # open workbook
            sheet = workbook.active

            # modify the desired cell
            print("chosen id - "+str(chosenid))
            print(chosenseats)
            print(busprice)
            if chosenid != -1 and chosenseats != -1:
                sheet["D"+str(chosenid+2)] = chosenseats + seats
                # save the file
                workbook.save(filename="data/seats.xlsx")
                return True
            else:
                df = pd.DataFrame([[date, source, destination, seats, bus, 0]], columns=[
                    "date", "source", "destination", "seats", "bus", "price"])
                book = load_workbook(f'{CURRENT_DIR}/data/seats.xlsx')
                writer = pd.ExcelWriter(
                    f'{CURRENT_DIR}/data/seats.xlsx', engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(
                    writer, startrow=writer.sheets['Sheet1'].max_row, index=False, header=False)
                writer.save()
                return True

        except Exception as e:
            print(e)
            return False

    # the function that authenticates the operator

    def operator_authenticator(self, phone, code):
        credentials = []
        try:
            excel_file = 'data/operators.xlsx'
            df = pd.read_excel(excel_file)
            res = df.where(df['phone'] == int(phone)).dropna()
            for (xname, xphone, xcode, xbus) in res.values:
                if xcode == int(code):
                    print([xname, xphone, xcode, xbus])
                    credentials.extend([xname, xphone, xcode, xbus])
        except Exception as e:
            print(e)
        return credentials

    def get_info_by_pnr(self, phone, pnr):
        credentials = []
        try:
            excel_file = 'data/bookings.xlsx'
            df = pd.read_excel(excel_file)
            res = df.where(df['tel'] == int(phone)).dropna()
            for (xno, xbookingdate, xfullname, xtel, xsource, xdestination, xtraveldate, xseats, xbus, xprice, xtotalprice, xbookingpnr) in res.values:
                if str(xbookingpnr) == str(pnr):
                    credentials = [xtraveldate, xfullname,
                                   xsource, xdestination, xbus]
                    return credentials
        except Exception as e:
            print(e)
        return credentials

    def check_pnr(self, phone, pnr):
        try:
            excel_file = 'data/bookings.xlsx'
            df = pd.read_excel(excel_file)
            res = df.where(df['tel'] == int(phone)).dropna()

            for (xno, xbookingdate, xfullname, xtel, xsource, xdestination, xtraveldate, xseats, xbus, xprice, xtotalprice, xbookingpnr) in res.values:
                if xbookingpnr != -1:
                    if str(xbookingpnr) == str(pnr):
                        return True
        except Exception as e:
            print(e)
        return False

    def NOT_IMPLEMENTED(self, update, ctx):
        user = update.message
        chat_id = update.effective_chat.id
        if chat_id not in self.messages:
            self.messages[chat_id] = "en"

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages

        keyboard = [
            [InlineKeyboardButton(messages["book_bus"],
                                  callback_data="book_bus")],
            [InlineKeyboardButton(messages["check_buses"],
                                  callback_data="check_buses")],
            [InlineKeyboardButton(messages["referal_link"],
                                  url="https://www.google.com")],
            [InlineKeyboardButton(messages["sell_seats"],
                                  callback_data="sell_seats")],
            [InlineKeyboardButton(
                messages["language_switcher"], callback_data="switch_language")],
        ]

        reply_markup = InlineKeyboardMarkup(keyboard)
        update.message.reply_text(messages["welcome_message"] % user.chat.first_name,
                                  reply_markup=reply_markup)

    def query_handler(self, update, ctx):
        query = update.callback_query
        query.answer()
        chat_id = update.effective_chat.id

        if self.messages[chat_id] == "en":
            messages = self.english_messages
        else:
            messages = self.amharic_messages

        if query.data == "book_bus":
            self.book_bus(update, ctx)

        elif query.data == "quest_passenger":
            self.passenger_menu(update, ctx)

        elif query.data == "quest_operator":
            self.operator_menu(update, ctx)

        elif query.data == "switch_language":  # DOne
            if self.messages[chat_id] == "am":
                ctx.bot.send_message(
                    chat_id, text=messages["lang_edited"] % "english")
            else:
                ctx.bot.send_message(
                    chat_id, text=messages["lang_edited"] % "amharic")
            self.load_data(self.messages[chat_id] == "en", True, chat_id)
        elif query.data == "back_main":
            self.main_menu(update, ctx)

    def main(self):
        print("BOT IS ONLINE")
        self.dispatcher.add_handler(CommandHandler("start", self.start))
        self.dispatcher.add_handler(CallbackQueryHandler(self.query_handler))
        self.dispatcher.add_handler(
            MessageHandler(Filters.photo, self.handle_image))
        self.dispatcher.add_handler(MessageHandler(
            Filters.document, self.handle_document))
        self.dispatcher.add_handler(
            MessageHandler(Filters.text, self.handle_text))
        self.dispatcher.add_error_handler(self.error)
        self.updater.start_polling()
        self.updater.idle()


if __name__ == "__main__":
    worker = GuzoBusBot(TOKEN)
    worker.main()
